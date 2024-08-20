#!/usr/bin/env python
# coding: utf-8
# In[1]:
# # Imports

import pandas as pd
import os
import SPACE2
from SPACE2 import reg_def

import seaborn as sns
import matplotlib.pyplot as plt

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import Color
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl import load_workbook


# In[0]:
def create_subfolders(folder_path, folders):
    """
    Create subfolders within a specified folder.

    Parameters:
    - folder_path (str): The path to the main folder.
    - folders (list): List of subfolder names to be created.

    Returns:
    None
    """
    for folder in folders:
        if not os.path.exists(folder):
            os.makedirs(os.path.join(folder_path, folder), exist_ok=True)

def organize_pdb(cluster_csv = None, 
                 pdb_folder = None, 
                 output_folder = None, 
                 cluster_column = 'Cluster', # decide if want to use original or adjusted clusters
                 pdb_naming_format = 'mHER_H3_LABELID_unique_fv_ISEQ#_igfold.pdb', 
                 labels_to_organize = 'all', #Replace with 'binders' or 'non binders' if only want to analyze one label type
                 binder_priorities_to_ignore = ['unclustered'], 
                 nonbinder_priorities_to_ignore = ['out_of_bounds', 'unclustered']
                 ):
    """
    Organize PDB files based on the provided cluster information. For a given .csv file, will create a new folder for each cluster indicated, organized by label type and priority status, and save a csv file containing the output folder as column name and the associated .pdb files for all sequences found in the cluster.

    Parameters:
    - cluster_csv (str): Path to the CSV file containing cluster information.
    - pdb_folder (str): Path to the folder containing PDB files.
    - output_folder (str): Path to the output folder for organizing PDB files.
    - cluster_column (str): Column name for clusters in the CSV file. Can be original 'Clusters' or 'adjusted_clusters'
    - pdb_naming_format (str): Naming format for PDB files.
    - labels_to_organize (str): Type of labels to organize ('all', 'binders', or 'non binders').
    - binder_priorities_to_ignore (list): List of binder priorities to ignore.
    - nonbinder_priorities_to_ignore (list): List of non-binder priorities to ignore.

    Returns:
    None
    """
    ### extract dataframe from csv file containing assigned clusterd
    cluster_df = pd.read_csv(cluster_csv)

    ### separate binders and non binders into separate dfs and add to empty dictionary
    dfs_to_analyze = {}
    if labels_to_organize == 'all' or labels_to_organize =='binders':
        binders_df = cluster_df[cluster_df['Labels'] == 'Binder']
        binders_df = binders_df[~binders_df['priority'].isin(binder_priorities_to_ignore)]
        dfs_to_analyze['binders']=binders_df

    if labels_to_organize == 'all' or labels_to_organize =='non binders':
        nonbinders_df = cluster_df[cluster_df['Labels'] == 'Non Binder']
        nonbinders_df = nonbinders_df[~nonbinders_df['priority'].isin(nonbinder_priorities_to_ignore)]
        dfs_to_analyze['non_binders']=nonbinders_df

    ### create creat binder/non binder subfolders in output folder
    # create_subfolders(output_folder, dfs_to_analyze.keys())

    ### iterate through dfs of selected labels to analyze
    cluster_pdb_files = {}
    file_name = cluster_csv.split('/')[-1]
    file_name = file_name.replace('.csv', '')
    file_name = file_name.replace('dbscanClusters', 'DBsc')

    for label, df in dfs_to_analyze.items():
        
        # Set label title specification for selected dataframe being analyzed
        label_subfolder = os.path.join(output_folder, label)
        label_id = 'AgPos' if label == 'binders' else 'AgNeg'
        pdb_filename = pdb_naming_format.replace('LABELID', label_id)

        # create empty dictionary to store output cluster subfolders and all pdb files in said folder
        for index, row in df.iterrows():
            iseq = row['iseq']
            priority = row['priority']
            cluster = row[cluster_column]

            # create final pdb file name based on iseq of entry
            new_pdb_filename = pdb_filename.replace('ISEQ#', str(iseq))
            pdb_file = os.path.join(pdb_folder, new_pdb_filename)

            # save pdb file path to dictionary where key is final cluster subfolder
            column_name = f'{priority}_{cluster}'

            if column_name not in cluster_pdb_files:
                cluster_pdb_files[column_name] = [pdb_file]
            else:
                cluster_pdb_files[column_name].append(pdb_file)

    df_cluster_pdb_files = pd.DataFrame.from_dict(cluster_pdb_files, orient='index').transpose()
    save_file = f'{output_folder}/{file_name}_ClstrPDBs.csv'
    df_cluster_pdb_files.to_csv(save_file, index = False)

    # for cluster_save_file, pdb_list in cluster_pdb_files.items():
    #     save_cluster_data = pd.DataFrame({cluster_save_file: pdb_list})
    #     save_cluster_data.to_csv(cluster_save_file, index = False)

def add_file_path(cell_value, file_path):
    return os.path.join(file_path, str(cell_value))

def add_file_path_to_csv_and_save(cluster_pdbs_file, path_to_pdbs):
    clusters_df = pd.read_csv(cluster_pdbs_file)
    adjusted_df = clusters_df.applymap(lambda x: add_file_path(x, path_to_pdbs))
    adjusted_df.to_csv(cluster_pdbs_file, index=False)  # Avoid saving DataFrame index
    print(adjusted_df.head())
# In[]:

def space2_prep_folders(csv_folder = None, 
                 pdb_folder = None, 
                 output_folder = None, 
                 cluster_column = 'Cluster', # decide if want to use original or adjusted clusters
                 pdb_naming_format = 'mHER_H3_LABELID_unique_fv_ISEQ#_igfold.pdb', 
                 labels_to_organize = 'all', #Replace with 'binders' or 'non binders' if only want to analyze one label type
                 binder_priorities_to_ignore = ['unclustered'], 
                 nonbinder_priorities_to_ignore = ['out_of_bounds', 'unclustered']
                 ):
    """
    Uses organize_pdb function to iterate through all cluster csv files found in a given folder and outputs the organized folders into the given output folder

    Parameters:
    - csv_folder (str): Path to the folder containing CSV files.
    - pdb_folder (str): Path to the folder containing PDB files.
    - output_folder (str): Path to the output folder for organizing PDB files.
    - cluster_column (str): Column name for clusters in the CSV file.
    - pdb_naming_format (str): Naming format for PDB files.
    - labels_to_organize (str): Type of labels to organize ('all', 'binders', or 'non binders').
    - binder_priorities_to_ignore (list): List of binder priorities to ignore.
    - nonbinder_priorities_to_ignore (list): List of non-binder priorities to ignore.

    Returns:
    None
    """
    # extract all csv file from input csv_folder
    csv_files = [file for file in os.listdir(csv_folder) if file.endswith(".csv")]
    
    # # create list of folders to make in output folder and 
    # csv_folders_to_make = [os.path.join(output_folder, file) for file in csv_files]
    # csv_folders_to_make = [file.replace('.csv', '') for file in csv_folders_to_make]

    csv_files = [os.path.join(csv_folder, file) for file in csv_files]

    # for folder in csv_folders_to_make:
    #     if not os.path.exists(folder):
    #             os.mkdir(folder)

    # folder_name = folder.split('/')[-1]

    for cluster_csv in csv_files:
            # file_name = cluster_csv.split('/')
            # file_name = file_name[-1]
            # file_name = file_name.replace('.csv', '')
            # if file_name == folder_name:
        organize_pdb(
                cluster_csv = cluster_csv, 
                pdb_folder = pdb_folder, 
                output_folder = output_folder, 
                cluster_column = cluster_column,
                pdb_naming_format = pdb_naming_format, 
                labels_to_organize = labels_to_organize,
                binder_priorities_to_ignore = binder_priorities_to_ignore, 
                nonbinder_priorities_to_ignore = nonbinder_priorities_to_ignore
                )

# In[]
def prep_all_folders_for_space2(replacements, 
                                pdb_folder, 
                                input_folder_format, 
                                output_folder_format, 
                                replace_in_format,
                                cluster_column = 'Cluster', 
                                pdb_naming_format = 'mHER_H3_LABELID_unique_fv_ISEQ#_igfold.pdb', 
                                labels_to_organize = 'all', 
                                binder_priorities_to_ignore = ['unclustered'], 
                                nonbinder_priorities_to_ignore = ['out_of_bounds', 'unclustered']):
    """
    Prepares all folders for SPACE2 using above space2_prep_folders and organize_pdb functions. 

    Parameters:
    - replacements (str): Text to replace.
    - pdb_folder (str): Path to the folder containing PDB files.
    - input_folder_format (str): Format of the input folder.
    - output_folder_format (str): Format of the output folder.
    - replace_in_format (str): Text to replace in the format.
    - cluster_column (str): Name of clusters column in the clusters CSV file to be used.
    - pdb_naming_format (str): Naming format for PDB files.
    - labels_to_organize (str): Type of labels to organize ('all', 'binders', or 'non binders').
    - binder_priorities_to_ignore (list): List of binder priorities to ignore.
    - nonbinder_priorities_to_ignore (list): List of non-binder priorities to ignore.

    Returns:
    None
    """
    for metric in replacements:
        input_folder = input_folder_format.replace(replace_in_format, metric)
        output_folder = output_folder_format.replace(replace_in_format, metric)

        if not os.path.exists(output_folder):
                os.mkdir(output_folder)
        
        space2_prep_folders(
                    csv_folder = input_folder, 
                    pdb_folder = pdb_folder, 
                    output_folder = output_folder, 
                    cluster_column = cluster_column, # decide if want to use original or adjusted clusters
                    pdb_naming_format = pdb_naming_format, 
                    labels_to_organize = labels_to_organize, #Replace with 'binders' or 'non binders' if only want to analyze one label type
                    binder_priorities_to_ignore = binder_priorities_to_ignore, 
                    nonbinder_priorities_to_ignore =nonbinder_priorities_to_ignore
                    )
        
# In []:
def SPACE2_clustering(cluster_pdbs_file, clusters_to_analyze, priorities_to_analyze, cdr_selection = ["CDR_all"], chain_selection = ['fw_all'], rmsd_threshold = 1.25, algorithm = "agglomerative", n_jobs = 1):

    selection = []
    anchor = []
    for i in cdr_selection:
        selection.append(reg_def[i])
    for i in chain_selection:
        anchor.append(reg_def[i])
    cdr_selection = selection
    chain_selection=anchor

    cluster_df = pd.read_csv(cluster_pdbs_file)

    all_clusters = cluster_df.columns.tolist()

    if 'all' in clusters_to_analyze or clusters_to_analyze == 'all':
        clusters_to_analyze = all_clusters

    all_structures_df = {}

    count = 1

    for cluster in clusters_to_analyze:
        priority = cluster.split('_')[0]
        pdb_files = cluster_df[cluster].tolist()
        pdb_files = [str(x) for x in pdb_files if isinstance(x, (str, bytes, os.PathLike))]

        if cluster in all_clusters and cluster in clusters_to_analyze:
            if priority in priorities_to_analyze or 'all' in priorities_to_analyze:
                
                print(f'\t\tCALCULATING CLUSTER #{count} OF {len(all_clusters)}')
                count += 1

                if algorithm == 'agglomerative':
                    structure_df = SPACE2.agglomerative_clustering(pdb_files, selection = cdr_selection, anchors=chain_selection, cutoff=rmsd_threshold, n_jobs = n_jobs)
                    all_structures_df[f'{algorithm}_{cluster}'] = structure_df

                if algorithm == 'greedy':
                    structure_df = SPACE2.greedy_clustering(pdb_files, selection = cdr_selection, anchors=chain_selection, cutoff=rmsd_threshold, n_jobs = n_jobs)  
                    all_structures_df[f'{algorithm}_{cluster}'] = structure_df

                if algorithm == 'both':
                    structure_df = SPACE2.agglomerative_clustering(pdb_files, selection = cdr_selection, anchors=chain_selection, cutoff=rmsd_threshold, n_jobs = n_jobs)
                    all_structures_df[f'{algorithm}_{cluster}'] = structure_df

                    structure_df = SPACE2.greedy_clustering(pdb_files, selection = cdr_selection, anchors=chain_selection, cutoff=rmsd_threshold, n_jobs = n_jobs)                       
                    all_structures_df[f'{algorithm}_{cluster}'] = structure_df

    return all_structures_df

# In[]:
def SPACE2_summary(all_structures_df):
    rows = []
    columns = ['cluster_name', 'total_antibodies', 'n_structural_clusters', 'structural_cluster', 'n_abs', 'n_abs%', 'cluster_by_length']

    for red_cluster, cluster_df in all_structures_df.items():
        if red_cluster != 'summary':
            n_antibodies = cluster_df.shape[0]
            n_str_clusters = cluster_df['cluster_by_rmsd'].nunique()
            cluster_by_length = cluster_df['cluster_by_length'].unique()
            str_clusters = cluster_df['cluster_by_rmsd'].value_counts().to_dict()
            
            for c_name, c_count in str_clusters.items():
                pc_cluster = round(((c_count / n_antibodies) * 100), 1)
                row = [red_cluster, n_antibodies, n_str_clusters, c_name, c_count, pc_cluster, cluster_by_length]
                rows.append(row)

    summary_df = pd.DataFrame(rows, columns = columns)

    all_structures_df['summary'] = summary_df

    return all_structures_df

def edit_structural_clusters_names(all_structures_df):
    for red_cluster, df in all_structures_df.items():
        if red_cluster != 'summary':
            df['ID'] = df['ID'].apply(lambda x: os.path.basename(x))
            df['cluster_by_rmsd'] = df['cluster_by_rmsd'].apply(lambda x: os.path.basename(x))
        if red_cluster == 'summary':
            df['structural_cluster'] = df['structural_cluster'].apply(lambda x: os.path.basename(x))

    return all_structures_df

def edit_xl_summary(xl_file):
    workbook = openpyxl.load_workbook(xl_file)

    sheet = workbook['summary']

    # Add a duplicate column for 'cluster_name' at the first position
    sheet.insert_cols(2)
    sheet.insert_cols(3)

    # Set the header for the new column
    priority_cell = sheet.cell(row=1, column=2, value='priority_level')
    priority_cell.font = Font(bold=True)
    label_cell=sheet.cell(row=1, column=3, value='label')
    label_cell.font = Font(bold=True)

    # Populate the new column by splitting values from 'cluster_name'
    cluster_name_column = sheet['A']
    for i in range(1, len(cluster_name_column)):
        cluster_name = cluster_name_column[i].value.split('_')
        priority_level = cluster_name[1]
        cluster_id = cluster_name[-1]
        label = 'Non Binder' if 'Non Binder' in cluster_id else 'Binder'

        sheet.cell(row=i+1, column=2, value=priority_level)
        sheet.cell(row=i+1, column=3, value=label)


    rules = [
        {'type': 'text_contains', 'criteria': 'high', 'column': 'B', 'format': PatternFill(start_color='C6EFCD', end_color='C6EFCD', fill_type='solid')},
        {'type': 'text_contains', 'criteria': 'med', 'column': 'B', 'format': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')},
        {'type': 'text_contains', 'criteria': 'low', 'column': 'B', 'format': PatternFill(start_color='FEC7CE', end_color='FEC7CE', fill_type='solid')},
        {'type': 'color_scale', 'column': 'E', 'format': ('C4D79B', 'FABF8F', 'B1A0CF')},
        {'type': 'color_scale', 'column': 'H', 'format': ('DA9694', 'FFFF00', '92D050')}
        # Add more rules as needed
    ]
    for rule in rules:
        rule_type = rule.get('type')
        criteria = rule.get('criteria', '')
        column = rule.get('column')
        format_style = rule.get('format')

        if rule_type == 'text_contains':
            for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):  # Assuming data starts from row 2
                column_cells = sheet[column]

                for cell in column_cells:
                    if criteria in cell.value:
                        cell.fill = format_style

        elif rule_type == 'color_scale':
            sheet.conditional_formatting.add(f'{column}2:{column}{sheet.max_row}', openpyxl.formatting.rule.ColorScaleRule(start_type='min', start_color=Color(rgb=format_style[0]), mid_type='percentile', mid_value=50, mid_color=Color(rgb=format_style[1]), end_type='max', end_color=Color(rgb=format_style[2])))


    # Add filter
    sheet.auto_filter.ref = sheet.dimensions

    # Save the workbook
    workbook.save(xl_file)


# edit_xl_summary(xl_file)

def edit_xl_summary_folder(xl_folder):
    xl_files = [file for file in os.listdir(xl_folder) if file.endswith(".xlsx")]
    xl_files = [os.path.join(xl_folder, file) for file in xl_files]

    for file in xl_files:
        edit_xl_summary(file)

def save_SPACE2_results(all_structural_dfs, output_file = None, shorten_antibody_names = 'y'):
    if shorten_antibody_names == 'y':
        all_structural_dfs = edit_structural_clusters_names(all_structural_dfs)

    with pd.ExcelWriter(output_file) as writer:

        all_structural_dfs['summary'].to_excel(writer, sheet_name='summary', index=False)

        for sheet_name, df in all_structural_dfs.items():
            if sheet_name != 'summary':
                df.to_excel(writer, sheet_name = sheet_name, index = False)

    edit_xl_summary(output_file)

def combine_all_summaries(structural_clusters_folder):
    # Open output excel sheet
    output_xl = openpyxl.Workbook()
    output_sheet = output_xl.create_sheet(title='all_summaries')

    bold_font = Font(bold = True)

    header_row = [
        'reduction_file',
        'cluster',
        'priority_level',
        'label',
        'total_antibodies',
        'n_structural_clusters',
        'structural_cluster',
        'n_abs',
        'n_abs%', 
        'cluster_by_length',
        'dimred_metric',
        'dimred_n_components', 
        'dimred_n_neighbors', 
        'dimred_min_dist', 
        'dimred_comp1', 
        'dimred_comp2', 
        'clustering_algorithm', 
        'clustering_eps', 
        'clustering_min_samples', 
        'str_model', 
        'str_algorithm', 
        'str_rmsd'
    ]

    for header in header_row:
        cell = output_sheet.cell(row=1, column=header_row.index(header) + 1, value=header)
        cell.font = bold_font

    xl_files = []

    # extract all excel files in subfolders
    for root, dirs, files in os.walk(structural_clusters_folder):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xls'):
                xl_files.append(os.path.join(root, file))


    for file in xl_files:

        file_name = file.split('/')[-1]

        if '.xlsx' in file_name:
            file_name = file_name.replace('.xlsx', '')
        if '.xls' in file_name:
            file_name = file_name.replace('.xls', '')

        # extract dimensionality reductions parameters
        filename = file_name.split('_')
        dim_red = filename[0]
        drd = filename[1].split('-')
        # print(drd)

        metric, n_components, n_neighbors, min_dist, comp1, comp2 = [drd[i] for i in range(1, len(drd))]

        n_components, n_neighbors, min_dist, comp1, comp2 = float(n_components), float(n_neighbors), float(min_dist), float(comp1), float(comp2)
        # print(metric, n_components, n_neighbors, min_dist, comp1, comp2)

        # extract dim red clustering parameters
        dr_clustering = filename[2].split('-')
        dr_algorithm, dr_eps, min_samples = [dr_clustering[i] for i in range(0, len(dr_clustering))]
        dr_eps, min_samples = float(dr_eps), float(min_samples)
        # print(dr_algorithm, dr_eps, min_samples)


        # extract structural clustering parameters
        str_method, str_algorithm, rmsd = filename[3], filename[4], float(filename[5])

        additional_data = [metric, n_components, n_neighbors, min_dist, comp1, comp2, dr_algorithm, dr_eps, min_samples, str_method, str_algorithm, rmsd]
        # print(str_method, str_algorithm, rmsd)

        # load summary sheet
        workbook = openpyxl.load_workbook(file)
        sheet = workbook['summary']

        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            new_row = [file_name]
            for cell in row:
                new_cell = openpyxl.cell.Cell(output_sheet, value=cell.value)
                new_row.append(new_cell)

            for data in additional_data:
                new_row.append(data)

            output_sheet.append(new_row)

    return output_xl



def edit_all_summaries_xl(output_xl, save_path):

    sheet = output_xl.active

    rules = [
    {'type': 'text_contains', 'criteria': 'high', 'column': 'C', 'format': PatternFill(start_color='C6EFCD', end_color='C6EFCD', fill_type='solid')},

    {'type': 'text_contains', 'criteria': 'med', 'column': 'C', 'format': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')},

    {'type': 'text_contains', 'criteria': 'low', 'column': 'C', 'format': PatternFill(start_color='FEC7CE', end_color='FEC7CE', fill_type='solid')},

    {'type': 'color_scale', 'column': 'F', 'format': ('C4D79B', 'FABF8F', 'B1A0CF')},
    # Add more rules as needed
    ]

    # Add standard color gradient rules to select columns
    standard_gradient_columns = ['E','H', 'I', 'L', 'M', 'N', 'R', 'S', 'V']
    for columns in standard_gradient_columns:
        rules.append({'type': 'color_scale', 'column': columns, 'format': ('DA9694', 'FFFF00', '92D050')})

    for rule in rules:
        rule_type = rule.get('type')
        criteria = rule.get('criteria', '')
        column = rule.get('column')
        format_style = rule.get('format')

        if rule_type == 'text_contains':
            for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):  # Assuming data starts from row 2
                column_cells = sheet[column]

                for cell in column_cells:
                    cell_value = str(cell.value)
                    if criteria in cell_value:
                        cell.fill = format_style

        elif rule_type == 'color_scale':
            sheet.conditional_formatting.add(f'{column}2:{column}{sheet.max_row}', openpyxl.formatting.rule.ColorScaleRule(start_type='min', start_color=Color(rgb=format_style[0]), mid_type='percentile', mid_value=50, mid_color=Color(rgb=format_style[1]), end_type='max', end_color=Color(rgb=format_style[2])))

    sheet.auto_filter.ref = sheet.dimensions
    output_xl.save(save_path)



def create_all_summaries_xl(structural_clusters_folder):
    save_path = os.path.join(structural_clusters_folder, 'all_summaries.xlsx')

    output_xl =combine_all_summaries(structural_clusters_folder)
    edit_all_summaries_xl(output_xl, save_path)

# In[]:
def SPACE2_folder_analysis(cluster_pdbs_folder, output_folder, clusters_to_analyze, priorities_to_analyze, cdr_selection, chain_selection, rmsd_threshold, algorithm, n_jobs, shorten_structural_cluster_names):
    csv_files = [file for file in os.listdir(cluster_pdbs_folder) if file.endswith(".csv")]
    csv_files = [os.path.join(cluster_pdbs_folder, file) for file in csv_files]

    print('Files to analyze:')
    for f in csv_files:
        print(f'\t{f}')    

    count = 1
    for file in csv_files:
        print(f'CALCULATING FILE #{count} of {len(csv_files)}')
        file_name = file.split('/')[-1]
        file_name = file_name.replace('.csv', '')
        file_name = file_name.replace('ClstrPDBs', 'SPACE2')
        file_name = f'{output_folder}/{file_name}_{algorithm}_{rmsd_threshold}.xlsx'


        all_structures_df = SPACE2_clustering(file, clusters_to_analyze, priorities_to_analyze,cdr_selection = cdr_selection, chain_selection = chain_selection, rmsd_threshold = rmsd_threshold, algorithm = algorithm, n_jobs = n_jobs)

        all_structures_df = SPACE2_summary(all_structures_df)

        save_SPACE2_results(all_structures_df, output_file = file_name, shorten_antibody_names=shorten_structural_cluster_names)

# In[]:
def space2_analyze_multiple_folders(replacements, input_folder_format, output_folder_format, replace_in_format,clusters_to_analyze = ['all'], priorities_to_analyze = ['all'], cdr_selection= ['CDRH3'], chain_selection = ['fwH'], rmsd_threshold = 1.25, algorithm = 'agglomerative', n_jobs =1,shorten_structural_cluster_names = 'y'):
        
    for metric in replacements:
        input_folder = input_folder_format.replace(replace_in_format, metric)
        output_folder = output_folder_format.replace(replace_in_format, metric)

        if not os.path.exists(output_folder):
                os.mkdir(output_folder)

        SPACE2_folder_analysis(input_folder, output_folder, clusters_to_analyze, priorities_to_analyze, cdr_selection, chain_selection, rmsd_threshold, algorithm, n_jobs, shorten_structural_cluster_names)

def pca_space2_full_summary_df(file):

    header_row = [
        'cluster_name',
        'component_1',
        'component_2',
        'label',
        'total_antibodies',
        'n_structural_clusters',
        'structural_cluster',
        'cluster_by_length',
        'n_abs',
        'n_abs%', 
        'str_model', 
        'str_algorithm', 
        'str_rmsd'
    ]
    input_df = pd.read_excel(file, sheet_name='summary', engine='openpyxl')

    full_summary_df = pd.DataFrame(columns = header_row)

    # extract dimensionality reductions parameters
    file_name = file.split('/')[-1]
    file_name = file_name.replace('.xlsx', '') if '.xlsx' in file_name else file_name.replace('.xl', '')
    file_name = file_name.split('_')

    cluster_model, cluster_alg, rmsd_val = file_name[4], file_name[5], file_name[6]

    rmsd_val = rmsd_val.replace('-', '.')


    for index, row in input_df.iterrows():
        cluster_name = row['cluster_name']
        tota_antibodies = row['total_antibodies']
        n_structural_clusters = row['n_structural_clusters']
        structural_cluster = row['structural_cluster']
        n_abs = row['n_abs']
        n_abspc = row['n_abs%']	
        cluster_by_length = row['cluster_by_length']

        pca_stats = cluster_name.split('_')
        sub_cluster = pca_stats[-1]
        label = 'Non Binder' if 'NB' in sub_cluster else 'Binder'

        pcas = pca_stats[-2]
        pcas = pcas.replace('PCA', '')
        pcas = pcas.split('-')
        pc1 = pcas[0]
        pc2 = pcas[1]

        new_row = [cluster_name, pc1, pc2, label, tota_antibodies, n_structural_clusters, structural_cluster, cluster_by_length, n_abs, n_abspc, cluster_model, cluster_alg, rmsd_val]

        full_summary_df.loc[len(full_summary_df)] = new_row

    return full_summary_df

def append_pca_full_summary_to_all_summaries(full_summary_df, folder):
    save_file = os.path.join(folder, 'all_summaries.xlsx')

    if not os.path.exists(save_file):
        full_summary_df.to_excel(save_file, index=False, sheet_name = 'all_summaries')

    else:
        book = load_workbook(save_file)

        writer = pd.ExcelWriter(save_file, engine='openpyxl')
        writer.book = book

        # Read the existing sheet into a DataFrame
        existing_df = pd.read_excel(save_file, sheet_name='all_summaries')

        # Append the new DataFrame to the existing one
        merged_df = pd.concat([existing_df, full_summary_df], ignore_index=True)

        # Write the merged DataFrame back to the Excel sheet
        merged_df.to_excel(writer, index=False, sheet_name='all_summaries')

        # Save the changes to the Excel file
        writer.save()
        writer.close()


def edit_pca_all_summaries_xl(workbook, save_path):
    # Get the active sheet
    sheet = workbook.active

    rules = []

    # Add standard color gradient rules to select columns
    standard_gradient_columns = ['E', 'F', 'H', 'I', 'J']
    for column in standard_gradient_columns:
        rules.append({'type': 'color_scale', 'column': column, 'format': ('DA9694', 'FFFF00', '92D050')})

    for rule in rules:
        rule_type = rule.get('type')
        column = rule.get('column')
        format_style = rule.get('format')

        if rule_type == 'color_scale':
            sheet.conditional_formatting.add(f'{column}2:{column}{sheet.max_row}', openpyxl.formatting.rule.ColorScaleRule(start_type='min', start_color=Color(rgb=format_style[0]), mid_type='percentile', mid_value=50, mid_color=Color(rgb=format_style[1]), end_type='max', end_color=Color(rgb=format_style[2])))

    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(save_path)


def create_pca_all_summaries(file):
    folder = file.split('/')[:-1]
    folder = '/'.join(folder)
    all_summaries_file = os.path.join(folder, 'all_summaries.xlsx')

    all_summaries_df_list = []

    df = pca_space2_full_summary_df(file)
    all_summaries_df_list = [df]

    for df in all_summaries_df_list:
        append_pca_full_summary_to_all_summaries(df, folder)

    # Load the Excel workbook
    workbook = load_workbook(all_summaries_file)

    # Call the function with the workbook object
    edit_pca_all_summaries_xl(workbook, all_summaries_file)

    

#In[]:
def all_summaries_overview(all_summaries_xl, pc_cutoff = 90):
    all_df= pd.read_excel(all_summaries_xl, sheet_name = 'all_summaries')

    binders_df = all_df[all_df['label'] == 'Binder']
    non_binders_df =all_df[all_df['label'] == 'Non Binder']
    all_dfs = {'binders': binders_df, 'non binders':non_binders_df}

    column_names = [
        'label', 'dimred_metric', 
        'n_high', 'high_%n_abs_avg', f'% of high clusters >= {pc_cutoff}%', 
        'n_med', 'med_%n_abs_avg',f'% of med clusters >= {pc_cutoff}%',
        'n_low', 'low_%n_abs_avg',f'% of low clusters >= {pc_cutoff}%',
        'n_out-of-bounds', 'out-of-bounds_%n_abs_avg',f'% of out-of-bounds clusters >= {pc_cutoff}%',
        'outlier_clusters'
        ]
    
    all_overviews ={}
    overview_df = pd.DataFrame(columns= column_names)

    for key, sub_df in all_dfs.items():
        dimred_metrics = sub_df['dimred_metric'].unique()

        for metric in dimred_metrics:
            metric_df = sub_df[sub_df['dimred_metric']== metric]
            
            high = metric_df[metric_df['priority_level'] == 'high']
            n_high = high.shape[0]
            n_high_avg = round(high['n_abs%'].sum()/n_high, 2)
            pc_high = round(((high[high['n_abs%']>=pc_cutoff].shape[0]/n_high) * 100), 2) if n_high != 0 else 0

            med = metric_df[metric_df['priority_level'] == 'med']
            n_med = med.shape[0]
            n_med_avg = round(med['n_abs%'].sum()/n_med, 2)
            pc_med = round(((med[med['n_abs%']>=pc_cutoff].shape[0]/n_med) * 100), 2)if n_med != 0 else 0

            low = metric_df[metric_df['priority_level'] == 'low']
            n_low = low.shape[0]
            n_low_avg = round(low['n_abs%'].sum()/n_low, 2)
            pc_low = round(((low[low['n_abs%']>=pc_cutoff].shape[0]/n_low) * 100), 2) if n_low != 0 else 0

            out_of_bounds = metric_df[metric_df['priority_level'] == 'out']
            n_out_of_bounds = out_of_bounds.shape[0]
            n_oob_avg = round(out_of_bounds['n_abs%'].sum()/n_out_of_bounds, 2)
            pc_out_of_bounds = round(((out_of_bounds[out_of_bounds['n_abs%']>=pc_cutoff].shape[0]/n_out_of_bounds) * 100), 2) if n_out_of_bounds != 0 else None

            n_outliers = metric_df[metric_df['n_abs'] == 1].shape[0]

            overview_row = {
                'dimred_metric' : metric,
                'label': key,
                'n_high': n_high,
                'high_%n_abs_avg': n_high_avg,
                f'% of high clusters >= {pc_cutoff}%': pc_high,

                'n_med': n_med,
                'med_%n_abs_avg': n_med_avg,
                f'% of med clusters >= {pc_cutoff}%': pc_med,

                'n_low':n_low,
                'low_%n_abs_avg': n_low_avg,
                f'% of low clusters >= {pc_cutoff}%': pc_low,

                'n_out-of-bounds': n_out_of_bounds,
                'out-of-bounds_%n_abs_avg':n_oob_avg,
                f'% of out-of-bounds clusters >= {pc_cutoff}%': pc_out_of_bounds,

                'n_outlier_clusters': n_outliers
            }

            overview_df.loc[len(overview_df)] = overview_row

        tot_high = sub_df[sub_df['priority_level'] == 'high']
        tot_n_high = tot_high.shape[0]
        tot_n_high_avg = round(tot_high['n_abs%'].sum()/tot_n_high, 2)
        tot_pc_high = round(((tot_high[tot_high['n_abs%']>=pc_cutoff].shape[0]/tot_n_high) * 100), 2) if n_high != 0 else 0

        tot_med = sub_df[sub_df['priority_level'] == 'med']
        tot_n_med = tot_med.shape[0]
        tot_n_med_avg = round(tot_med['n_abs%'].sum()/tot_n_med, 2)
        tot_pc_med = round(((tot_med[tot_med['n_abs%']>=pc_cutoff].shape[0]/tot_n_med) * 100), 2)if n_med != 0 else 0

        tot_low = sub_df[sub_df['priority_level'] == 'low']
        tot_n_low = tot_low.shape[0]
        tot_n_low_avg = round(tot_low['n_abs%'].sum()/tot_n_low, 2)
        tot_pc_low = round(((tot_low[tot_low['n_abs%']>=pc_cutoff].shape[0]/tot_n_low) * 100), 2) if n_low != 0 else 0

        tot_out_of_bounds = sub_df[sub_df['priority_level'] == 'out']
        tot_n_out_of_bounds = tot_out_of_bounds.shape[0]
        tot_n_oob_avg = round(tot_out_of_bounds['n_abs%'].sum()/tot_n_out_of_bounds, 2)
        tot_pc_out_of_bounds = round(((tot_out_of_bounds[tot_out_of_bounds['n_abs%']>=pc_cutoff].shape[0]/tot_n_out_of_bounds) * 100), 2) if n_out_of_bounds != 0 else 0

        tot_n_outliers = sub_df[sub_df['n_abs'] == 1].shape[0]

        total_row = {
            'dimred_metric' : f'all {key}',
            'label': key,
            'n_high': tot_n_high,
            'high_%n_abs_avg': tot_n_high_avg,
            f'% of high clusters >= {pc_cutoff}%': tot_pc_high,
            'n_med': tot_n_med,
            'med_%n_abs_avg': tot_n_med_avg,
            f'% of med clusters >= {pc_cutoff}%': tot_pc_med,
            'n_low':tot_n_low,
            'low_%n_abs_avg': tot_n_low_avg,
            f'% of low clusters >= {pc_cutoff}%': tot_pc_low,
            'n_out-of-bounds': tot_n_out_of_bounds,
            'out-of-bounds_%n_abs_avg':tot_n_oob_avg,
            f'% of out-of-bounds clusters >= {pc_cutoff}%': tot_pc_out_of_bounds,
            'outlier_clusters': tot_n_outliers
            }


        overview_df.loc[len(overview_df)] = total_row
        all_overviews[key] = overview_df

    with pd.ExcelWriter(all_summaries_xl, engine='openpyxl', mode = 'a') as writer:

        overview_df.to_excel(writer, sheet_name=f'{pc_cutoff}%_cutoff_overview', index=False)

        # Open the workbook and the new sheet
        workbook = writer.book
        sheet = workbook[f'{pc_cutoff}%_cutoff_overview']
        sheet.auto_filter.ref = sheet.dimensions


        # Apply color scale rule to each column (excluding the first column containing 'dimred_metric')
        for col_idx in range(2, sheet.max_column + 1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            range_string = f'{column_letter}2:{column_letter}{sheet.max_row}'

            color_scale_rule = ColorScaleRule(
                start_type='min', start_color='DA9694',
                mid_type='percentile', mid_value=50, mid_color='FFFF00',
                end_type='max', end_color='92D050'
            )

            sheet.conditional_formatting.add(range_string, color_scale_rule)

    return overview_df

#In[]:
def summaries_violin_plot(x_axes, y_axes, data_set, title, font_size, save_path, inner_plot='box'):
    # Create a violin plot
    sns.violinplot(x=x_axes, y=y_axes, data=data_set, inner=inner_plot)
    sns.swarmplot()

    # Set the title with the optimal font size
    plt.title(title, fontsize=font_size)

    plt.tight_layout()
    # Show the plot
    plt.savefig(os.path.join(save_path, title))

    plt.show()